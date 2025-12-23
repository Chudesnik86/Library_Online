"""
Excel file parser for importing books
"""
from typing import List, Dict, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def parse_books_excel(file_path: str) -> Tuple[List[Dict], List[str]]:
    """
    Parse Excel file and extract book data
    
    Expected Excel format:
    - First row: Headers (Title, Author, ISBN, Category, Total Copies, Available Copies)
    - Following rows: Book data
    
    Args:
        file_path: Path to Excel file
    
    Returns:
        Tuple of (books_data: List[Dict], errors: List[str])
    """
    books_data = []
    errors = []
    
    workbook = None
    try:
        # Open workbook for reading
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        # Find header row (first row with data)
        header_row = None
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
            values = [cell.value for cell in row if cell.value]
            if any(val and isinstance(val, str) and val.lower() in ['title', 'название', 'название книги'] for val in values):
                header_row = row_idx
                break
        
        if not header_row:
            errors.append("Не найдена строка заголовков. Убедитесь, что первая строка содержит заголовки.")
            if workbook:
                workbook.close()
            return books_data, errors
        
        # Read headers
        headers = {}
        header_cells = sheet[header_row]
        for col_idx, cell in enumerate(header_cells, start=1):
            if cell.value:
                header_text = str(cell.value).strip().lower()
                # Map various header names to standard fields
                if any(x in header_text for x in ['title', 'название', 'название книги', 'книга']):
                    headers['title'] = col_idx
                elif any(x in header_text for x in ['author', 'автор', 'автор книги']):
                    headers['author'] = col_idx
                elif any(x in header_text for x in ['isbn']):
                    headers['isbn'] = col_idx
                elif any(x in header_text for x in ['category', 'категория', 'жанр', 'genre']):
                    headers['category'] = col_idx
                elif any(x in header_text for x in ['description', 'описание', 'описание книги']):
                    headers['description'] = col_idx
                elif any(x in header_text for x in ['cover', 'обложка', 'cover image', 'изображение', 'image', 'url']):
                    headers['cover_image'] = col_idx
                elif any(x in header_text for x in ['total', 'всего', 'total copies', 'всего экземпляров']):
                    headers['total_copies'] = col_idx
                elif any(x in header_text for x in ['available', 'доступно', 'available copies', 'доступно экземпляров']):
                    headers['available_copies'] = col_idx
        
        if 'title' not in headers:
            errors.append("Не найдена колонка с названием книги. Убедитесь, что в файле есть колонка 'Title' или 'Название'.")
            if workbook:
                workbook.close()
            return books_data, errors
        
        # Read data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=False), start=header_row + 1):
            # Skip empty rows
            if not any(cell.value for cell in row):
                continue
            
            book_data = {}
            
            # Extract title (required)
            title_cell = row[headers['title'] - 1]
            title = str(title_cell.value).strip() if title_cell.value else None
            
            if not title or title.lower() in ['none', 'null', '']:
                errors.append(f"Строка {row_idx}: Отсутствует название книги")
                continue
            
            book_data['title'] = title
            
            # Extract author (optional)
            if 'author' in headers:
                author_cell = row[headers['author'] - 1]
                author = str(author_cell.value).strip() if author_cell.value else None
                if author and author.lower() not in ['none', 'null', '']:
                    book_data['author'] = author
            
            # Extract ISBN (optional)
            if 'isbn' in headers:
                isbn_cell = row[headers['isbn'] - 1]
                isbn = str(isbn_cell.value).strip() if isbn_cell.value else None
                if isbn and isbn.lower() not in ['none', 'null', '']:
                    book_data['isbn'] = isbn
            
            # Extract category (optional)
            if 'category' in headers:
                category_cell = row[headers['category'] - 1]
                category = str(category_cell.value).strip() if category_cell.value else None
                if category and category.lower() not in ['none', 'null', '']:
                    book_data['category'] = category
            
            # Extract description (optional)
            if 'description' in headers:
                desc_cell = row[headers['description'] - 1]
                description = str(desc_cell.value).strip() if desc_cell.value else None
                if description and description.lower() not in ['none', 'null', '']:
                    book_data['description'] = description
            
            # Extract cover image URL (optional)
            if 'cover_image' in headers:
                cover_cell = row[headers['cover_image'] - 1]
                cover_image = str(cover_cell.value).strip() if cover_cell.value else None
                if cover_image and cover_image.lower() not in ['none', 'null', '']:
                    book_data['cover_image'] = cover_image
            
            # Extract total copies (optional, default 1)
            if 'total_copies' in headers:
                total_cell = row[headers['total_copies'] - 1]
                try:
                    total = int(total_cell.value) if total_cell.value else 1
                    book_data['total_copies'] = max(1, total)
                except (ValueError, TypeError):
                    book_data['total_copies'] = 1
            else:
                book_data['total_copies'] = 1
            
            # Extract available copies (optional, default = total_copies)
            if 'available_copies' in headers:
                available_cell = row[headers['available_copies'] - 1]
                try:
                    available = int(available_cell.value) if available_cell.value else book_data['total_copies']
                    book_data['available_copies'] = max(0, min(available, book_data['total_copies']))
                except (ValueError, TypeError):
                    book_data['available_copies'] = book_data['total_copies']
            else:
                book_data['available_copies'] = book_data['total_copies']
            
            books_data.append(book_data)
        
        if not books_data:
            errors.append("Не найдено ни одной книги в файле. Убедитесь, что данные начинаются со второй строки.")
        
    except Exception as e:
        errors.append(f"Ошибка при чтении файла: {str(e)}")
    finally:
        # Always close the workbook to release the file lock
        if workbook:
            try:
                workbook.close()
            except Exception:
                pass  # Ignore errors when closing
    
    return books_data, errors


def parse_customers_excel(file_path: str) -> Tuple[List[Dict], List[str]]:
    """
    Parse Excel file and extract customer data
    
    Expected Excel format:
    - First row: Headers (ID, Name, Address, Zip, City, Phone, Email)
    - Following rows: Customer data
    
    Args:
        file_path: Path to Excel file
    
    Returns:
        Tuple of (customers_data: List[Dict], errors: List[str])
    """
    customers_data = []
    errors = []
    
    workbook = None
    try:
        # Open workbook for reading
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        # Find header row (first row with data)
        header_row = None
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
            values = [cell.value for cell in row if cell.value]
            if any(val and isinstance(val, str) and val.lower() in ['name', 'имя', 'фио', 'fio', 'customer', 'читатель'] for val in values):
                header_row = row_idx
                break
        
        if not header_row:
            errors.append("Не найдена строка заголовков. Убедитесь, что первая строка содержит заголовки.")
            if workbook:
                workbook.close()
            return customers_data, errors
        
        # Read headers
        headers = {}
        header_cells = sheet[header_row]
        for col_idx, cell in enumerate(header_cells, start=1):
            if cell.value:
                header_text = str(cell.value).strip().lower()
                # Map various header names to standard fields
                if any(x in header_text for x in ['id', 'идентификатор', 'номер']):
                    headers['id'] = col_idx
                elif any(x in header_text for x in ['name', 'имя', 'фио', 'fio', 'customer', 'читатель', 'название']):
                    headers['name'] = col_idx
                elif any(x in header_text for x in ['address', 'адрес', 'адрес проживания']):
                    headers['address'] = col_idx
                elif any(x in header_text for x in ['zip', 'индекс', 'почтовый индекс', 'postal', 'postcode']):
                    headers['zip'] = col_idx
                elif any(x in header_text for x in ['city', 'город', 'населенный пункт']):
                    headers['city'] = col_idx
                elif any(x in header_text for x in ['phone', 'телефон', 'телефонный номер', 'tel']):
                    headers['phone'] = col_idx
                elif any(x in header_text for x in ['email', 'почта', 'электронная почта', 'e-mail']):
                    headers['email'] = col_idx
        
        if 'name' not in headers:
            errors.append("Не найдена колонка с именем читателя. Убедитесь, что в файле есть колонка 'Name' или 'Имя'.")
            if workbook:
                workbook.close()
            return customers_data, errors
        
        # Read data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=False), start=header_row + 1):
            # Skip empty rows
            if not any(cell.value for cell in row):
                continue
            
            customer_data = {}
            
            # Extract name (required)
            name_cell = row[headers['name'] - 1]
            name = str(name_cell.value).strip() if name_cell.value else None
            
            if not name or name.lower() in ['none', 'null', '']:
                errors.append(f"Строка {row_idx}: Отсутствует имя читателя")
                continue
            
            customer_data['name'] = name
            
            # Extract ID (optional, will be generated if not provided)
            if 'id' in headers:
                id_cell = row[headers['id'] - 1]
                customer_id = str(id_cell.value).strip() if id_cell.value else None
                if customer_id and customer_id.lower() not in ['none', 'null', '']:
                    customer_data['id'] = customer_id
            
            # Extract address (optional)
            if 'address' in headers:
                address_cell = row[headers['address'] - 1]
                address = str(address_cell.value).strip() if address_cell.value else None
                if address and address.lower() not in ['none', 'null', '']:
                    customer_data['address'] = address
            
            # Extract zip (optional)
            if 'zip' in headers:
                zip_cell = row[headers['zip'] - 1]
                try:
                    zip_code = int(zip_cell.value) if zip_cell.value else None
                    if zip_code:
                        customer_data['zip'] = zip_code
                except (ValueError, TypeError):
                    pass  # Skip invalid zip codes
            
            # Extract city (optional)
            if 'city' in headers:
                city_cell = row[headers['city'] - 1]
                city = str(city_cell.value).strip() if city_cell.value else None
                if city and city.lower() not in ['none', 'null', '']:
                    customer_data['city'] = city
            
            # Extract phone (optional)
            if 'phone' in headers:
                phone_cell = row[headers['phone'] - 1]
                phone = str(phone_cell.value).strip() if phone_cell.value else None
                if phone and phone.lower() not in ['none', 'null', '']:
                    customer_data['phone'] = phone
            
            # Extract email (optional)
            if 'email' in headers:
                email_cell = row[headers['email'] - 1]
                email = str(email_cell.value).strip() if email_cell.value else None
                if email and email.lower() not in ['none', 'null', '']:
                    customer_data['email'] = email
            
            customers_data.append(customer_data)
        
        if not customers_data:
            errors.append("Не найдено ни одного читателя в файле. Убедитесь, что данные начинаются со второй строки.")
        
    except Exception as e:
        errors.append(f"Ошибка при чтении файла: {str(e)}")
    finally:
        # Always close the workbook to release the file lock
        if workbook:
            try:
                workbook.close()
            except Exception:
                pass  # Ignore errors when closing
    
    return customers_data, errors


def parse_issues_excel(file_path: str) -> Tuple[List[Dict], List[str]]:
    """
    Parse Excel file and extract issue (book loan) data
    
    Expected Excel format:
    - First row: Headers (Book ID, Book, Customer ID, Customer, Date of issue, Return date)
    - Following rows: Issue data
    
    Args:
        file_path: Path to Excel file
    
    Returns:
        Tuple of (issues_data: List[Dict], errors: List[str])
    """
    issues_data = []
    errors = []
    
    workbook = None
    try:
        # Open workbook for reading
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        # Find header row (first row with data)
        header_row = None
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
            values = [cell.value for cell in row if cell.value]
            if any(val and isinstance(val, str) and any(x in val.lower() for x in ['book id', 'book_id', 'книга', 'id книги', 'date of issue', 'date_issue', 'дата выдачи']) for val in values):
                header_row = row_idx
                break
        
        if not header_row:
            errors.append("Не найдена строка заголовков. Убедитесь, что первая строка содержит заголовки.")
            if workbook:
                workbook.close()
            return issues_data, errors
        
        # Read headers
        headers = {}
        header_cells = sheet[header_row]
        for col_idx, cell in enumerate(header_cells, start=1):
            if cell.value:
                header_text = str(cell.value).strip().lower()
                # Map various header names to standard fields
                if any(x in header_text for x in ['book id', 'book_id', 'id книги', 'книга id']):
                    headers['book_id'] = col_idx
                elif any(x in header_text for x in ['book', 'книга', 'название книги', 'book title', 'book_title']):
                    headers['book'] = col_idx
                elif any(x in header_text for x in ['customer id', 'customer_id', 'id читателя', 'читатель id', 'client id']):
                    headers['customer_id'] = col_idx
                elif any(x in header_text for x in ['customer', 'читатель', 'клиент', 'customer name', 'customer_name', 'client']):
                    headers['customer'] = col_idx
                elif any(x in header_text for x in ['date of issue', 'date_issue', 'дата выдачи', 'issued', 'выдано', 'issue date']):
                    headers['date_issued'] = col_idx
                elif any(x in header_text for x in ['return date', 'return_date', 'дата возврата', 'returned', 'возвращено', 'date return']):
                    headers['date_return'] = col_idx
        
        if 'book_id' not in headers:
            errors.append("Не найдена колонка с ID книги (Book ID). Убедитесь, что в файле есть колонка 'Book ID'.")
            if workbook:
                workbook.close()
            return issues_data, errors
        
        if 'customer_id' not in headers:
            errors.append("Не найдена колонка с ID читателя (Customer ID). Убедитесь, что в файле есть колонка 'Customer ID'.")
            if workbook:
                workbook.close()
            return issues_data, errors
        
        if 'date_issued' not in headers:
            errors.append("Не найдена колонка с датой выдачи (Date of issue). Убедитесь, что в файле есть колонка 'Date of issue'.")
            if workbook:
                workbook.close()
            return issues_data, errors
        
        # Read data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=False), start=header_row + 1):
            # Skip empty rows
            if not any(cell.value for cell in row):
                continue
            
            issue_data = {}
            
            # Extract book_id (required)
            book_id_cell = row[headers['book_id'] - 1]
            book_id = str(book_id_cell.value).strip() if book_id_cell.value else None
            
            if not book_id or book_id.lower() in ['none', 'null', '']:
                errors.append(f"Строка {row_idx}: Отсутствует ID книги")
                continue
            
            issue_data['book_id'] = book_id
            
            # Extract book title (optional, for validation)
            if 'book' in headers:
                book_cell = row[headers['book'] - 1]
                book_title = str(book_cell.value).strip() if book_cell.value else None
                if book_title and book_title.lower() not in ['none', 'null', '']:
                    issue_data['book_title'] = book_title
            
            # Extract customer_id (required)
            customer_id_cell = row[headers['customer_id'] - 1]
            customer_id = str(customer_id_cell.value).strip() if customer_id_cell.value else None
            
            if not customer_id or customer_id.lower() in ['none', 'null', '']:
                errors.append(f"Строка {row_idx}: Отсутствует ID читателя")
                continue
            
            issue_data['customer_id'] = customer_id
            
            # Extract customer name (optional, for validation)
            if 'customer' in headers:
                customer_cell = row[headers['customer'] - 1]
                customer_name = str(customer_cell.value).strip() if customer_cell.value else None
                if customer_name and customer_name.lower() not in ['none', 'null', '']:
                    issue_data['customer_name'] = customer_name
            
            # Extract date_issued (required)
            date_issued_cell = row[headers['date_issued'] - 1]
            date_issued = date_issued_cell.value
            
            if not date_issued:
                errors.append(f"Строка {row_idx}: Отсутствует дата выдачи")
                continue
            
            # Convert date to string format YYYY-MM-DD
            try:
                if isinstance(date_issued, str):
                    # Try to parse string date
                    from datetime import datetime
                    date_issued = datetime.strptime(date_issued, '%Y-%m-%d').date()
                elif hasattr(date_issued, 'date'):
                    # Excel date object
                    date_issued = date_issued.date()
                elif hasattr(date_issued, 'isoformat'):
                    # Already a date object
                    pass
                else:
                    # Try to convert to date
                    from datetime import datetime
                    date_issued = datetime.strptime(str(date_issued), '%Y-%m-%d').date()
                
                issue_data['date_issued'] = date_issued.isoformat()
            except Exception as e:
                errors.append(f"Строка {row_idx}: Неверный формат даты выдачи: {str(e)}")
                continue
            
            # Extract date_return (optional)
            if 'date_return' in headers:
                date_return_cell = row[headers['date_return'] - 1]
                date_return = date_return_cell.value
                
                if date_return:
                    try:
                        if isinstance(date_return, str):
                            from datetime import datetime
                            date_return = datetime.strptime(date_return, '%Y-%m-%d').date()
                        elif hasattr(date_return, 'date'):
                            date_return = date_return.date()
                        elif hasattr(date_return, 'isoformat'):
                            pass
                        else:
                            from datetime import datetime
                            date_return = datetime.strptime(str(date_return), '%Y-%m-%d').date()
                        
                        issue_data['date_return'] = date_return.isoformat()
                        issue_data['status'] = 'returned'
                    except Exception as e:
                        errors.append(f"Строка {row_idx}: Неверный формат даты возврата: {str(e)}")
                        # Continue without return date
                        issue_data['status'] = 'issued'
                else:
                    issue_data['status'] = 'issued'
            else:
                issue_data['status'] = 'issued'
            
            issues_data.append(issue_data)
        
        if not issues_data:
            errors.append("Не найдено ни одной выдачи в файле. Убедитесь, что данные начинаются со второй строки.")
        
    except Exception as e:
        errors.append(f"Ошибка при чтении файла: {str(e)}")
    finally:
        # Always close the workbook to release the file lock
        if workbook:
            try:
                workbook.close()
            except Exception:
                pass  # Ignore errors when closing
    
    return issues_data, errors

